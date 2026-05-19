"""
FlexFlow - Gerador de Arquivo ONET Mock para Testes de Produção
Gera um arquivo Excel profissional com 50 linhas e estrutura completa de 22 campos
Inclui campos financeiros: Vl.Unit, Total Item, e Valor Total do Pedido
"""

import pandas as pd
from datetime import datetime, timedelta
from decimal import Decimal
import sys
import io
import random

# Fix encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def generate_onet_mock():
    """
    Gera arquivo Excel mock com as 22 colunas ONET completas para produção.
    
    Estrutura de 22 campos:
    1. Pedido
    2. Cliente
    3. SKU
    4. Descrição
    5. Qtd
    6. Unidade
    7. Largura
    8. Comprimento
    9. Lead Time
    10. Data Entrega
    11. Data Faturamento
    12. % ICMS
    13. Bloqueio (Crédito)
    14. Saldo
    15. Atraso
    16. Condição Pagamento
    17. Frete
    18. Vendedor
    19. IPI
    20. Vl.Unit (Valor Unitário)
    21. Total Item (Qtd * Vl.Unit)
    22. Valor Total do Pedido (Soma de todos Total Item do PO)
    
    Gera 50 linhas com:
    - Mix de SKUs existentes e novos
    - Mix de status de crédito (bloqueado/liberado)
    - Mix de tipos de pedido (Replacement/Normal)
    - Variação realista de quantidades, dimensões e valores
    - Integridade financeira: Total Item = Qtd * Vl.Unit
    - Integridade de PO: Valor Total do Pedido = Soma dos Total Item
    """
    
    # Data base para cálculos
    today = datetime.now()
    
    # SKUs existentes no sistema (baseado em material_costs)
    existing_skus = [
        'PP-1000', 'ABS-2000', 'PE-1000', 'PET-1000', 'PC-1000',
        'PS-1000', 'PVC-1000', 'PMMA-1000', 'PA-1000', 'POM-1000'
    ]
    
    # SKUs novos (para testar warnings)
    new_skus = [
        'PAB-035', 'PAB-036', 'PAB-037', 'PAB-038', 'PAB-039',
        'NEW-100', 'NEW-101', 'NEW-102', 'NEW-103', 'NEW-104'
    ]
    
    # Clientes variados
    clients = [
        'Indústria Automotiva XYZ Ltda',
        'Embalagens Premium S.A.',
        'Eletrônicos Delta Corp',
        'Móveis Modernos Ltda',
        'Farmacêutica BioHealth',
        'Cosméticos Bella Vita',
        'Alimentos Natureza S.A.',
        'Tecnologia Inovare Ltda',
        'Construção Forte & Cia',
        'Têxtil Fashion Group'
    ]
    
    # Vendedores
    sellers = [
        'João Silva',
        'Maria Santos',
        'Pedro Oliveira',
        'Ana Costa',
        'Carlos Ferreira'
    ]
    
    # Condições de pagamento
    payment_terms = [
        '30 dias',
        '45 dias',
        '60 dias',
        '30/60 dias',
        'À vista',
        '15/30/45 dias'
    ]
    
    # Descrições base
    descriptions = [
        'Tampa Reservatório {} Natural',
        'Caixa Organizadora {} com Tampa',
        'Painel Frontal {} Customizado',
        'Gaveta Modular {} com Divisórias',
        'Frasco {} Cristal 500ml',
        'Pote {} Transparente com Lacre',
        'Bandeja {} Retangular',
        'Suporte {} para Eletrônicos',
        'Protetor {} Industrial',
        'Componente {} Técnico'
    ]
    
    data = []
    po_totals = {}  # Track total value per PO
    po_clients = {}  # Track client per PO (CONSISTENCY FIX)
    po_sellers = {}  # Track seller per PO (CONSISTENCY FIX)
    po_payment_terms = {}  # Track payment terms per PO (CONSISTENCY FIX)
    
    # Gerar 50 linhas com variação realista
    for i in range(50):
        # Determinar se é SKU existente ou novo (70% existente, 30% novo)
        is_existing = random.random() < 0.7
        sku = random.choice(existing_skus) if is_existing else random.choice(new_skus)
        
        # Extrair material do SKU para descrição
        material = sku.split('-')[0]
        description = random.choice(descriptions).format(material)
        
        # Adicionar flag de personalização em alguns casos
        is_personalized = random.random() < 0.3
        if is_personalized:
            description += ' - PERSONALIZADO'
        
        # Determinar se é Replacement (20% dos casos)
        is_replacement = random.random() < 0.2
        
        # Bloqueio de crédito (15% dos casos)
        is_blocked = random.random() < 0.15
        credit_status = 'BLOQUEADO' if is_blocked else 'LIBERADO'
        
        # Atraso (10% dos casos, apenas para pedidos não-replacement)
        has_delay = random.random() < 0.1 and not is_replacement
        delay_days = random.randint(1, 15) if has_delay else 0
        
        # Calcular datas
        lead_time = random.randint(10, 45)
        delivery_date = today + timedelta(days=lead_time + random.randint(0, 10))
        invoice_date = delivery_date - timedelta(days=random.randint(1, 3))
        
        # Ajustar data de entrega se houver atraso
        if has_delay:
            delivery_date = delivery_date - timedelta(days=delay_days)
        
        # Gerar valores realistas
        quantity = random.choice([50, 100, 150, 200, 250, 300, 500, 750, 1000, 1500, 2000])
        width = round(random.uniform(50.0, 500.0), 1)
        length = round(random.uniform(50.0, 500.0), 1)
        icms_rate = random.choice([0.0, 7.0, 12.0, 18.0])
        freight = round(random.uniform(100.0, 2000.0), 2)
        ipi_rate = random.choice([0.0, 5.0, 10.0, 15.0])
        
        # FINANCIAL FIELDS - NEW
        # Valor unitário realista (R$ 5.00 a R$ 150.00)
        unit_value = round(random.uniform(5.0, 150.0), 2)
        
        # Total do item = Qtd * Vl.Unit (INTEGRITY CHECK)
        item_total = round(quantity * unit_value, 2)
        
        # Saldo (para Replacement, pode ter saldo pendente)
        balance = 0
        if is_replacement:
            balance = random.randint(0, quantity // 2)
        
        # Número do pedido - Agrupar alguns itens no mesmo PO
        po_type = 'REP' if is_replacement else 'ONET'
        # Create 10 different POs, so multiple items per PO
        po_base = 1001 + (i // 5)  # Every 5 items share a PO
        po_number = f'{po_type}-2026-{po_base}'
        
        # Track PO totals for Valor Total do Pedido
        if po_number not in po_totals:
            po_totals[po_number] = 0.0
            # CONSISTENCY FIX: Assign client, seller, and payment terms once per PO
            po_clients[po_number] = random.choice(clients)
            po_sellers[po_number] = random.choice(sellers)
            po_payment_terms[po_number] = random.choice(payment_terms)
        po_totals[po_number] += item_total
        
        # Use consistent client, seller, and payment terms for this PO
        client_name = po_clients[po_number]
        seller_name = po_sellers[po_number]
        payment_term = po_payment_terms[po_number]
        
        row = {
            'Pedido': po_number,
            'Cliente': client_name,  # CONSISTENCY FIX: Use PO-specific client
            'SKU': sku,
            'Descrição': description,
            'Qtd': quantity,
            'Unidade': 'UN',
            'Largura': width,
            'Comprimento': length,
            'Lead Time': lead_time,
            'Data Entrega': delivery_date.strftime('%d/%m/%Y'),
            'Data Faturamento': invoice_date.strftime('%d/%m/%Y'),
            '% ICMS': icms_rate,
            'Bloqueio': credit_status,
            'Saldo': balance,
            'Atraso': delay_days,
            'Condição Pagamento': payment_term,  # CONSISTENCY FIX: Use PO-specific payment term
            'Frete': freight,
            'Vendedor': seller_name,  # CONSISTENCY FIX: Use PO-specific seller
            'IPI': ipi_rate,
            'Vl.Unit': unit_value,
            'Total Item': item_total,
            'Valor Total do Pedido': 0.0  # Will be filled in next pass
        }
        
        data.append(row)
    
    # Second pass: Fill in Valor Total do Pedido for each row
    for row in data:
        row['Valor Total do Pedido'] = po_totals[row['Pedido']]
    
    # Criar DataFrame
    df = pd.DataFrame(data)
    
    # Definir ordem exata das colunas (22 campos)
    columns_order = [
        'Pedido', 'Cliente', 'SKU', 'Descrição', 'Qtd', 'Unidade',
        'Largura', 'Comprimento', 'Lead Time', 'Data Entrega',
        'Data Faturamento', '% ICMS', 'Bloqueio', 'Saldo', 'Atraso',
        'Condição Pagamento', 'Frete', 'Vendedor', 'IPI',
        'Vl.Unit', 'Total Item', 'Valor Total do Pedido'
    ]
    
    df = df[columns_order]
    
    # Gerar arquivo Excel com formatação profissional
    output_file = 'onet_production_test_50_rows.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Pedidos ONET', index=False)
        
        # Obter worksheet para formatação
        worksheet = writer.sheets['Pedidos ONET']
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 18,  # Pedido
            'B': 35,  # Cliente
            'C': 12,  # SKU
            'D': 50,  # Descrição
            'E': 8,   # Qtd
            'F': 10,  # Unidade
            'G': 12,  # Largura
            'H': 14,  # Comprimento
            'I': 12,  # Lead Time
            'J': 15,  # Data Entrega
            'K': 18,  # Data Faturamento
            'L': 10,  # % ICMS
            'M': 12,  # Bloqueio
            'N': 10,  # Saldo
            'O': 10,  # Atraso
            'P': 20,  # Condição Pagamento
            'Q': 12,  # Frete
            'R': 18,  # Vendedor
            'S': 10,  # IPI
            'T': 12,  # Vl.Unit
            'U': 14,  # Total Item
            'V': 20   # Valor Total do Pedido
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatar cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Alinhar dados e formatar valores monetários
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for idx, cell in enumerate(row):
                if idx in [4, 6, 7, 8, 11, 13, 14, 18]:  # Colunas numéricas (não monetárias)
                    cell.alignment = Alignment(horizontal='right')
                elif idx in [16, 19, 20, 21]:  # Colunas monetárias (Frete, Vl.Unit, Total Item, Valor Total)
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = 'R$ #,##0.00'
                elif idx in [9, 10]:  # Datas
                    cell.alignment = Alignment(horizontal='center')
                elif idx == 12:  # Bloqueio (status)
                    cell.alignment = Alignment(horizontal='center')
                    # Colorir células bloqueadas
                    if cell.value == 'BLOQUEADO':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
                else:
                    cell.alignment = Alignment(horizontal='left')
    
    # Estatísticas do arquivo gerado
    existing_count = len(df[df['SKU'].isin(existing_skus)])
    new_count = len(df[df['SKU'].isin(new_skus)])
    blocked_count = len(df[df['Bloqueio'] == 'BLOQUEADO'])
    replacement_count = len(df[df['Pedido'].str.startswith('REP')])
    personalized_count = len(df[df['Descrição'].str.contains('PERSONALIZADO')])
    delayed_count = len(df[df['Atraso'] > 0])
    unique_pos = df['Pedido'].nunique()
    
    # Financial statistics
    total_items_value = df['Total Item'].sum()
    avg_unit_value = df['Vl.Unit'].mean()
    
    print("=" * 80)
    print("[OK] Arquivo ONET Mock de Produção Gerado com Sucesso!")
    print("=" * 80)
    print(f"\nArquivo: {output_file}")
    print(f"Total de linhas: {len(df)}")
    print(f"Total de campos: 22")
    print(f"Total de POs únicos: {unique_pos}")
    print("\n📊 Estatísticas dos Dados:")
    print(f"   ✅ SKUs EXISTENTES: {existing_count} ({existing_count/len(df)*100:.1f}%)")
    print(f"   ⚠️  SKUs NOVOS: {new_count} ({new_count/len(df)*100:.1f}%)")
    print(f"   🔒 Crédito BLOQUEADO: {blocked_count} ({blocked_count/len(df)*100:.1f}%)")
    print(f"   🔄 Pedidos REPLACEMENT: {replacement_count} ({replacement_count/len(df)*100:.1f}%)")
    print(f"   🎨 Itens PERSONALIZADOS: {personalized_count} ({personalized_count/len(df)*100:.1f}%)")
    print(f"   ⏰ Pedidos com ATRASO: {delayed_count} ({delayed_count/len(df)*100:.1f}%)")
    
    print("\n💰 Estatísticas Financeiras:")
    print(f"   💵 Valor Total dos Itens: R$ {total_items_value:,.2f}")
    print(f"   📊 Valor Unitário Médio: R$ {avg_unit_value:,.2f}")
    print(f"   🎯 Integridade: Total Item = Qtd × Vl.Unit ✓")
    print(f"   🎯 Integridade: Valor Total do Pedido = Σ Total Item ✓")
    
    print("\n📋 Estrutura de 22 Campos:")
    for i, col in enumerate(columns_order, 1):
        marker = "💰" if col in ['Vl.Unit', 'Total Item', 'Valor Total do Pedido'] else "  "
        print(f"   {marker} {i:2d}. {col}")
    
    print("\n🎯 Casos de Teste Incluídos:")
    print("   [x] SKUs existentes em material_costs")
    print("   [x] SKUs novos para testar warning de custo não encontrado")
    print("   [x] Itens personalizados (requerem toggle 'Personalizado')")
    print("   [x] Pedidos Replacement com saldo pendente")
    print("   [x] Bloqueios de crédito (status BLOQUEADO)")
    print("   [x] Pedidos com atraso na entrega")
    print("   [x] Diferentes clientes e vendedores")
    print("   [x] Variação de ICMS, IPI e condições de pagamento")
    print("   [x] Variação de quantidades (50 a 2000 unidades)")
    print("   [x] Variação de dimensões (50mm a 500mm)")
    print("   [x] Valores unitários realistas (R$ 5.00 a R$ 150.00)")
    print("   [x] Integridade financeira: Total Item = Qtd × Vl.Unit")
    print("   [x] Integridade de PO: Valor Total = Σ Total Item")
    print("   [x] Múltiplos itens por PO (agrupamento realista)")
    
    print("\n" + "=" * 80)
    print("\n🚀 Próximos Passos:")
    print("   1. Importe este arquivo na Mesa de Conferência (Staging Area)")
    print("   2. Verifique se todos os 22 campos são reconhecidos")
    print("   3. Valide a integridade financeira (Total Item e Valor Total do Pedido)")
    print("   4. Teste o validador de integridade de PO")
    print("   5. Valide os cálculos de Margem (CM) com valores reais")
    print("   6. Confirme que paginação/virtual scrolling funciona")
    print("=" * 80)
    
    # Gerar também um arquivo de resumo
    summary_file = 'onet_production_test_summary.txt'
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("FLEXFLOW - RESUMO DO ARQUIVO DE TESTE DE PRODUÇÃO (22 CAMPOS)\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Arquivo gerado: {output_file}\n")
        f.write(f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Total de linhas: {len(df)}\n")
        f.write(f"Total de campos: 22\n")
        f.write(f"Total de POs únicos: {unique_pos}\n\n")
        f.write("ESTATÍSTICAS:\n")
        f.write(f"  - SKUs existentes: {existing_count}\n")
        f.write(f"  - SKUs novos: {new_count}\n")
        f.write(f"  - Crédito bloqueado: {blocked_count}\n")
        f.write(f"  - Pedidos Replacement: {replacement_count}\n")
        f.write(f"  - Itens personalizados: {personalized_count}\n")
        f.write(f"  - Pedidos com atraso: {delayed_count}\n\n")
        f.write("ESTATÍSTICAS FINANCEIRAS:\n")
        f.write(f"  - Valor Total dos Itens: R$ {total_items_value:,.2f}\n")
        f.write(f"  - Valor Unitário Médio: R$ {avg_unit_value:,.2f}\n\n")
        f.write("CAMPOS (22):\n")
        for i, col in enumerate(columns_order, 1):
            marker = "[NOVO]" if col in ['Vl.Unit', 'Total Item', 'Valor Total do Pedido'] else ""
            f.write(f"  {i:2d}. {col} {marker}\n")
        f.write("\nINTEGRIDADE FINANCEIRA:\n")
        f.write("  ✓ Total Item = Qtd × Vl.Unit (validado)\n")
        f.write("  ✓ Valor Total do Pedido = Σ Total Item por PO (validado)\n")
    
    print(f"\n📄 Resumo salvo em: {summary_file}")
    
    # Verify integrity for a sample PO
    print("\n🔍 Verificação de Integridade (Amostra):")
    sample_po = df['Pedido'].iloc[0]
    po_items = df[df['Pedido'] == sample_po]
    calculated_total = po_items['Total Item'].sum()
    declared_total = po_items['Valor Total do Pedido'].iloc[0]
    print(f"   PO: {sample_po}")
    print(f"   Itens: {len(po_items)}")
    print(f"   Soma calculada: R$ {calculated_total:,.2f}")
    print(f"   Valor declarado: R$ {declared_total:,.2f}")
    print(f"   Diferença: R$ {abs(calculated_total - declared_total):.2f}")
    print(f"   Status: {'✓ ÍNTEGRO' if abs(calculated_total - declared_total) < 0.01 else '✗ DIVERGENTE'}")

if __name__ == "__main__":
    generate_onet_mock()
