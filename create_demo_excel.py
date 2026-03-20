"""
Script para criar planilha demo do FlexFlow para o Kickoff
"""
import sys
import codecs

# Fix encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from decimal import Decimal

# Criar workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pedidos Demo"

# Cabeçalhos
headers = ['Pedido', 'Cliente', 'SKU', 'Qtd', 'Vlr Unit', 'Custo MP', 'Custo MO', 'Custo Energia', 'Custo Gas', 'Personalizado']
ws.append(headers)

# Estilizar cabeçalho
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dados realistas com margens lucrativas
# Margem típica: 20-35%
dados = [
    # PO-001: Cliente Acme Corp - 2 itens (1 personalizado)
    ['PO-2024-001', 'Acme Corp', 'FLEX-1000', 100, 150.00, 80.00, 25.00, 8.00, 5.00, 'Sim'],
    ['PO-2024-001', 'Acme Corp', 'FLEX-2000', 50, 280.00, 150.00, 45.00, 12.00, 8.00, 'Não'],
    
    # PO-002: Cliente Beta Industries - 2 itens (1 personalizado)
    ['PO-2024-002', 'Beta Industries', 'FLEX-3000', 75, 420.00, 220.00, 70.00, 18.00, 12.00, 'Não'],
    ['PO-2024-002', 'Beta Industries', 'FLEX-CUSTOM-A', 25, 650.00, 350.00, 110.00, 25.00, 15.00, 'Sim'],
    
    # PO-003: Cliente Gamma Solutions - 2 itens (1 personalizado)
    ['PO-2024-003', 'Gamma Solutions', 'FLEX-4000', 120, 195.00, 105.00, 32.00, 10.00, 6.00, 'Não'],
    ['PO-2024-003', 'Gamma Solutions', 'FLEX-CUSTOM-B', 40, 580.00, 310.00, 95.00, 22.00, 13.00, 'Sim'],
    
    # PO-004: Cliente Delta Corp - 2 itens
    ['PO-2024-004', 'Delta Corp', 'FLEX-5000', 90, 325.00, 175.00, 52.00, 15.00, 10.00, 'Não'],
    ['PO-2024-004', 'Delta Corp', 'FLEX-6000', 60, 480.00, 260.00, 78.00, 20.00, 14.00, 'Não'],
]

# Adicionar dados
for row in dados:
    ws.append(row)

# Ajustar largura das colunas
column_widths = {
    'A': 15,  # Pedido
    'B': 20,  # Cliente
    'C': 18,  # SKU
    'D': 8,   # Qtd
    'E': 12,  # Vlr Unit
    'F': 12,  # Custo MP
    'G': 12,  # Custo MO
    'H': 15,  # Custo Energia
    'I': 12,  # Custo Gas
    'J': 15,  # Personalizado
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Formatar valores monetários
for row in ws.iter_rows(min_row=2, min_col=5, max_col=9):
    for cell in row:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = 'R$ #,##0.00'

# Centralizar quantidade
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

# Salvar
wb.save('planilha_demo_promaflex.xlsx')
print("✅ Planilha 'planilha_demo_promaflex.xlsx' criada com sucesso!")
print(f"📊 Total de linhas: {len(dados)} itens em 4 pedidos")
print(f"🎨 Itens personalizados: 3 (marcados como 'Sim')")
